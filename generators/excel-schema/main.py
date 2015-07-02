#!/usr/bin/env python3
import os
import csv
import openpyxl
import sys

OPENPYXL_MAJOR_VERSION = int(openpyxl.__version__[0])
DATETIME_FORMAT = '%Y-%m-%dT%H:%M:%S%Z'
# Minimum possible header.
HEADER = (
    'domain',
    'entity',
    'attribute',
    'value',
)


def generate(file_name, domain):
    # Creating a CSV writer that feeds to stdout.
    w = csv.writer(sys.stdout)

    # Write the header to start.
    w.writerow(HEADER)

    wb = openpyxl.load_workbook(file_name, use_iterators=True)
    sheets = wb.get_sheet_names()

    for i, sheet_name in enumerate(sheets):
        sheet = wb.get_sheet_by_name(sheet_name)
        column_names = _column_names(wb, sheet_name)
        rows = sheet.iter_rows()

        # Skip header row because we already collected the column names.
        header = next(rows)

        for j, cell in enumerate(header):
            label = list(column_names)[j]
            column = cell.column
            w.writerow([
                domain,
                label,
                'label',
                label
            ])
            w.writerow([
                domain,
                label,
                'sheet',
                sheet.title
            ])
            w.writerow([
                domain,
                label,
                'column',
                column
            ])


def _column_names(workbook, sheet_name):
    sheet = workbook.get_sheet_by_name(sheet_name)
    first_row = next(sheet.iter_rows())

    # Renamed attribute in version 2+
    if OPENPYXL_MAJOR_VERSION > 1:
        return [c.value for c in first_row]

    return [c.internal_value for c in first_row]


def main():
    usage = """Origins Excel Schema Generator.
    Usage:
        excel <file_name> [--domain=<domain>]

    Options:
        --domain=<domain>  The domain for new facts.
        -h --help  Show usage info.
    """ #noqa

    from docopt import docopt

    # Trim './main.py excel' from the line.
    args = docopt(usage, version='0.1')

    # Collect arg values
    file_name = os.path.join(os.getcwd(), args['<file_name>'])
    domain = args['--domain'] + '.schema'
    domain = domain.replace(' ', '_').lower()

    # No need to validate input because docopt makes sure all
    # arguments are given
    generate(file_name, domain)

if __name__ == '__main__':
    main()
