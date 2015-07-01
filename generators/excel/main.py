#!/usr/bin/env python3
import os
import csv
import openpyxl
import sys

OPENPYXL_MAJOR_VERSION = int(openpyxl.__version__[0])
DATETIME_FORMAT = '%Y-%m-%dT%H:%M:%S%Z'
# Minimum possible header.
HEADER = (
    'domain',
    'entity',
    'attribute',
    'value',
)


def generate(file_name, domain, identifier):
    # Creating a CSV writer that feeds to stdout.
    w = csv.writer(sys.stdout)

    # Write the header to start.
    w.writerow(HEADER)

    wb = openpyxl.load_workbook(file_name, use_iterators=True)
    sheets = wb.get_sheet_names()

    for i, sheet_name in enumerate(sheets):
        sheet = wb.get_sheet_by_name(sheet_name)
        if domain is not None:
            sheet_domain = domain + '.' + sheet_name.lower().replace(' ', '_')
        else:
            sheet_domain = sheet_name.lower().replace(' ', '_')
        column_names = _column_names(wb, sheet_name)
        rows = sheet.iter_rows()

        # Skip header row because we already collected the column names.
        next(rows)
        rows = list(rows)

        for j, row in enumerate(rows):
            row = rows[j]

            if identifier != []:
                row_id = create_identifier(
                    row,
                    sheet_name,
                    identifier,
                    column_names
                )
            else:
                row_id = j

            for k, cell in enumerate(row):
                if cell.value is not None:
                    w.writerow([
                        sheet_domain,
                        row_id,
                        column_names[k],
                        cell.value,
                    ])


def _column_names(workbook, sheet_name):
    sheet = workbook.get_sheet_by_name(sheet_name)
    first_row = next(sheet.iter_rows())

    # Renamed attribute in version 2+
    if OPENPYXL_MAJOR_VERSION > 1:
        return [c.value for c in first_row]

    return [c.internal_value for c in first_row]


def create_identifier(line, sheet, identifier_columns, column_names):
    id = ''
    line = list(line)
    for column in identifier_columns:
        if sheet == column.split(':')[0]:
            id = id + str(
                line[column_names.index(column.split(':')[1])].value
            ) + '_'
    id = id[0:len(id)-1]
    return id


def main():
    usage = """Origins Excel Generator.
    Usage:
        excel <file_name> [--domain=<domain>] [--identifier=<identfier>...]

    Options:
        --domain=<domain>  The domain for new facts.
        --identifier=<identifier>...  The column name(s) used as the primary
            key for each sheet. Provide the sheet name, followed by ':',
            followed by the column name. Ex: --identifier="Sheet1:Artist"
        -h --help  Show usage info.
    """ #noqa

    from docopt import docopt

    # Trim './main.py excel' from the line.
    args = docopt(usage, version='0.1')

    # Collect arg values
    file_name = os.path.join(os.getcwd(), args['<file_name>'])
    domain = None
    identifier = None

    if args['--identifier'] is not None:
        identifier = args['--identifier']
        for id in identifier:
            # Make sure the user
            if len(id.split(':')) is not 2:
                print('To list identifiers, provide sheet AND colum name.')
                print('For example: "Sheet1:Track_Id"')
                sys.exit()

    if args['--domain'] is not None:
        domain = args['--domain']

    # No need to validate input because docopt makes sure all
    # arguments are given
    print(identifier)
    generate(file_name, domain, identifier)

if __name__ == '__main__':
    main()
